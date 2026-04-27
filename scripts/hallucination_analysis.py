#!/usr/bin/env python3
"""Deterministic hallucination analysis for Nightmare Extraction Test.

Ground truth JSON carries the authoritative universe of values for each
document (header fields, raw_ground_truth.fields, locations, claims,
drivers, etc.). A hallucination here is an extracted value that doesn't
land anywhere in that universe even after normalization.

No LLMs. No source-document parsing. Just JSON walking + string/number
matching.

Usage:
    python scripts/hallucination_analysis.py \\
        --ground-truth ground_truth/ \\
        --results results/ \\
        --output results/hallucination_report.json
"""

import argparse
import json
import re
import subprocess
from collections import defaultdict
from pathlib import Path
from typing import Any


# ── value normalization ────────────────────────────────────────────────

_PUNCT_RE = re.compile(r"[^\w\s]")
_WS_RE = re.compile(r"\s+")


def norm_string(s: Any) -> str:
    if s is None:
        return ""
    out = str(s).strip().lower()
    out = _PUNCT_RE.sub(" ", out)
    out = _WS_RE.sub(" ", out).strip()
    return out


def as_float(v: Any) -> float | None:
    if v is None or v is True or v is False:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "").replace("$", "").replace(" ", "")
    s = s.rstrip("%")
    try:
        return float(s)
    except ValueError:
        return None


def looks_meaningful_string(s: str) -> bool:
    """Filter out things that aren't worth flagging (boilerplate, nulls)."""
    if not s:
        return False
    if s in {"null", "none", "n a", "not applicable", "na", "unknown",
             "tbd", "pending", "open", "closed", "yes", "no", "true",
             "false", "active", "inactive"}:
        return False
    # One-char or single-word generic types aren't worth flagging
    if len(s) <= 2:
        return False
    return True


# ── Universe builder ──────────────────────────────────────────────────


def collect_universe(gt_doc: dict) -> tuple[set, list]:
    """Walk ground truth for a single doc; return (string_set, number_list)."""
    strings: set[str] = set()
    numbers: list[float] = []

    def walk(obj: Any):
        if isinstance(obj, dict):
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for v in obj:
                walk(v)
        elif isinstance(obj, (int, float)) and not isinstance(obj, bool):
            numbers.append(float(obj))
            # Also add string form so digit-for-digit matches work both ways
            strings.add(norm_string(obj))
        elif isinstance(obj, str):
            ns = norm_string(obj)
            if ns:
                strings.add(ns)
            # If the string is a number, capture it as a number too
            f = as_float(obj)
            if f is not None:
                numbers.append(f)

    walk(gt_doc)
    return strings, numbers


# ── Rich source-document universe ─────────────────────────────────────
#
# The benchmark GT JSON only models a subset of the fields that appear on
# the rendered document. Anything the model extracts that IS on the page
# but NOT in the GT would get flagged as a hallucination. To avoid these
# false positives, we also ingest:
#   - Every generator *_truth*.json next to the packet (richer than GT)
#   - pdftotext of every source PDF
#   - All cell values from every xlsx
#   - Raw text of every csv


def _is_content_key(k: str) -> bool:
    """Heuristic: ingest dict keys that look like human-readable content
    labels ("Decking", "Roof Age", "SIC Code") and skip plumbing keys
    ("fields", "value", "doc_type", "locations", "insured_name").

    A content label has an uppercase letter or a space; plumbing keys
    are all lowercase snake_case.
    """
    if not isinstance(k, str) or len(k) < 3:
        return False
    return any(c.isupper() for c in k) or " " in k


def _ingest_value(obj: Any, strings: set[str], numbers: list[float]) -> None:
    if isinstance(obj, dict):
        for k, v in obj.items():
            if _is_content_key(k):
                ns = norm_string(k)
                if ns:
                    strings.add(ns)
            _ingest_value(v, strings, numbers)
    elif isinstance(obj, list):
        for v in obj:
            _ingest_value(v, strings, numbers)
    elif isinstance(obj, (int, float)) and not isinstance(obj, bool):
        numbers.append(float(obj))
        strings.add(norm_string(obj))
    elif isinstance(obj, str):
        ns = norm_string(obj)
        if ns:
            strings.add(ns)
        f = as_float(obj)
        if f is not None:
            numbers.append(f)


_OCR_READER = None  # lazy-init easyocr to avoid startup cost when not needed


def _ocr_pdf(path: Path) -> str:
    """OCR a PDF that has no text layer (N2 scanned docs)."""
    global _OCR_READER
    cache = path.with_suffix(path.suffix + ".ocr.txt")
    if cache.exists():
        try:
            return cache.read_text()
        except Exception:
            pass
    # No cache → need live OCR. pdf2image + easyocr are required here;
    # silently returning "" on ImportError was a real footgun (2026-04-18:
    # a venv missing easyocr produced N2-N5 universes with no text layer,
    # inflating fabrication rates by several points across all models and
    # masquerading as an analyzer drift for hours before we caught it).
    try:
        import pdf2image
        import numpy as np
        import easyocr
    except ImportError as e:
        raise RuntimeError(
            f"OCR fallback needed for {path.name} (no text layer, no cache), "
            f"but a dependency is missing: {e}. Install easyocr + pdf2image "
            f"(see requirements.txt). Rates are environment-dependent without OCR."
        )
    if _OCR_READER is None:
        _OCR_READER = easyocr.Reader(["en"], gpu=False, verbose=False)
    try:
        pages = pdf2image.convert_from_path(str(path), dpi=150)
    except Exception:
        return ""
    chunks = []
    for page in pages:
        try:
            result = _OCR_READER.readtext(np.array(page), detail=0, paragraph=False)
            chunks.extend(result)
        except Exception:
            continue
    text = "\n".join(chunks)
    try:
        cache.write_text(text)
    except Exception:
        pass
    return text


def _ingest_text(text: str, strings: set[str], numbers: list[float]) -> None:
    # Token-level ingestion only. Line-level ingestion was dropped in the
    # 2026-04-17 audit: combined with substring matching (also removed) it
    # inflated the universe so much that almost any fabricated value could
    # be cleared as "found in source." With exact/compact matching, a line
    # like "Total premium: $1,234,567" contributes little beyond its tokens
    # ("total", "premium", "1234567") - which we already add below.
    for tok in text.split():
        tok = tok.strip(".,;:$%()[]{}'\"")
        ns = norm_string(tok)
        if ns:
            strings.add(ns)
            # Also index each sub-word. norm_string collapses internal
            # punctuation to whitespace, so a raw token like "273-7259" or
            # "Lender's" produces a multi-word normed form. Without this,
            # the composed-string matcher can never find the sub-word
            # components ("273", "7259", "lender"), and the analyzer
            # false-flags real identifier/compound values as fabrications.
            # Gated at len>=2 to match the matcher's single-char filter.
            if " " in ns:
                for sub in ns.split():
                    if len(sub) >= 2:
                        strings.add(sub)
        f = as_float(tok)
        if f is not None:
            numbers.append(f)


def _ingest_pdf(path: Path, strings: set[str], numbers: list[float]) -> None:
    pdftotext_failed = False
    try:
        out = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"],
            capture_output=True, text=True, timeout=20,
        ).stdout
    except FileNotFoundError:
        # pdftotext not installed. This is environment-critical - without it,
        # the universe is limited to GT JSON + generator source-of-truth only,
        # which dramatically deflates or inflates hallucination rates
        # depending on what's missing. Crash loudly so this is caught at
        # CI/launch time, not discovered from weird numbers after publish.
        raise RuntimeError(
            "pdftotext binary is required for hallucination analysis "
            "(install poppler-utils). Aborting rather than silently "
            "producing environment-dependent numbers."
        )
    except Exception:
        pdftotext_failed = True
        out = ""
    if not out.strip():
        # No text layer - fall back to OCR. Necessary for N2 scanned docs
        # and some N3/N4 PDFs with invisible or corrupted text layers.
        ocr_out = _ocr_pdf(path)
        if not ocr_out.strip():
            _PDF_INGEST_FAILURES.append(str(path))
        out = ocr_out
    elif pdftotext_failed:
        _PDF_INGEST_FAILURES.append(str(path))
    _ingest_text(out, strings, numbers)


_PDF_INGEST_FAILURES: list[str] = []


def _ingest_xlsx(path: Path, strings: set[str], numbers: list[float]) -> None:
    try:
        import openpyxl
    except ImportError:
        return
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return
    for sh in wb.sheetnames:
        try:
            ws = wb[sh]
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None:
                        continue
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        numbers.append(float(v))
                        strings.add(norm_string(v))
                    else:
                        ns = norm_string(v)
                        if ns:
                            strings.add(ns)
                        f = as_float(v)
                        if f is not None:
                            numbers.append(f)
        except Exception:
            continue


def _ingest_csv(path: Path, strings: set[str], numbers: list[float]) -> None:
    try:
        text = path.read_text(errors="ignore")
    except Exception:
        return
    for line in text.splitlines():
        for cell in line.split(","):
            cell = cell.strip().strip('"').strip("'")
            ns = norm_string(cell)
            if ns:
                strings.add(ns)
            f = as_float(cell)
            if f is not None:
                numbers.append(f)


def _derive_packet_dir(gt_data: dict, gt_file: Path | None = None) -> Path | None:
    """Given a loaded packet GT dict, find the generator packet dir.

    document_path looks like
      /...generator.../output/nightmare/N1_easy/doc_70001/documents/acord_140_70001.pdf
    (maintainer workspace, absolute) or
      documents/acord_140_70001.pdf
    (public repo, relative to gt_file). The packet dir is the
    grandparent of the PDF path either way.
    """
    docs = gt_data.get("documents", {}) or {}
    for doc_gt in docs.values():
        p = doc_gt.get("document_path")
        if not p:
            continue
        pp = Path(p)
        if not pp.is_absolute() and gt_file is not None:
            pp = gt_file.resolve().parent / pp
        # .../doc_70001/documents/acord_140_70001.pdf -> .../doc_70001
        if pp.parent.name == "documents":
            return pp.parent.parent
    return None


def build_packet_universe(gt_data: dict, gt_file: Path | None = None) -> tuple[set[str], list[float]]:
    """Rich packet universe: packet GT + generator GT JSONs + rendered docs."""
    strings: set[str] = set()
    numbers: list[float] = []

    # (1) Packet GT for every doc in the packet
    for doc_gt in gt_data.get("documents", {}).values():
        s, n = collect_universe(doc_gt)
        strings |= s
        numbers.extend(n)

    # (2) Generator source-of-truth artifacts
    vdir = _derive_packet_dir(gt_data, gt_file)
    if vdir is not None and vdir.is_dir():
        gdir = vdir / "ground_truth"
        if gdir.is_dir():
            for jf in gdir.glob("*.json"):
                try:
                    _ingest_value(json.loads(jf.read_text()), strings, numbers)
                except Exception:
                    continue

        # (3) Rendered documents (PDF / xlsx / csv)
        ddir = vdir / "documents"
        if ddir.is_dir():
            for f in ddir.iterdir():
                suf = f.suffix.lower()
                if suf == ".pdf":
                    _ingest_pdf(f, strings, numbers)
                elif suf in (".xlsx", ".xlsm"):
                    _ingest_xlsx(f, strings, numbers)
                elif suf == ".csv":
                    _ingest_csv(f, strings, numbers)

    return strings, numbers


def _compact(s: str) -> str:
    """Strip all non-alphanumerics for identifier-style comparisons."""
    return re.sub(r"[^a-z0-9]", "", s.lower())


_UNIVERSE_COMPACT_CACHE: dict[int, set[str]] = {}
_UNIVERSE_TOKEN_CACHE: dict[int, set[str]] = {}


def _universe_compact_set(universe: set[str]) -> set[str]:
    """Memoized compact form of the universe for identifier-style matches."""
    key = id(universe)
    cached = _UNIVERSE_COMPACT_CACHE.get(key)
    if cached is not None:
        return cached
    compacted = {_compact(u) for u in universe}
    compacted.discard("")
    _UNIVERSE_COMPACT_CACHE[key] = compacted
    return compacted


def _universe_token_set(universe: set[str]) -> set[str]:
    """Memoized set of whitespace-separated tokens from universe entries,
    filtering out tokens shorter than 3 chars to avoid matching on "f",
    "x", etc. Used by composed-string acceptance so that a value like
    "Ford F-350" matches when the universe stores {make: "Ford",
    model: "F-350"} - "350" appears as a token of the universe entry
    "f 350" (the normalized form of "F-350").
    """
    key = id(universe)
    cached = _UNIVERSE_TOKEN_CACHE.get(key)
    if cached is not None:
        return cached
    tokens: set[str] = set()
    for entry in universe:
        for t in entry.split():
            if len(t) >= 3:
                tokens.add(t)
    _UNIVERSE_TOKEN_CACHE[key] = tokens
    return tokens


def string_in_universe(val: str, universe: set[str]) -> bool:
    """Match the normalized value against the universe.

    Strategy (three tiers, cheapest first):
      1. Exact normalized match.
      2. Compact-form exact match (so "CL-2023-12345" matches
         "CL202312345" after punctuation/whitespace collapse). Requires
         ≥4 compact chars to avoid rubber-stamping trivially short values.
      3. Composed-string acceptance: for values with ≥2 meaningful tokens,
         if ≥80% of tokens individually appear in the universe, treat as
         real. Catches cases where the model concatenates real source
         values into one field (e.g., "LOC-001: Preston Center Tower -
         8117 Preston Road, Dallas, TX 75225 (Occupancy: Office)") - the
         components are all in the source, the combined form isn't.
         Also catches short 2-word source phrases like "Wet Pipe" or
         "District 6" that don't show up as a combined universe entry.
         For 2-token values the 80% threshold rounds to 2/2 - both
         tokens must match - which is the right conservatism at that
         length.

    Substring matching was removed in the 2026-04-17 audit: it was
    rubber-stamping fabricated values as "not hallucinated" whenever they
    happened to share a substring with any long universe string, and
    systematically deflated published hallucination rates.
    """
    if not val:
        return False
    if val in universe:
        return True
    val_c = _compact(val)
    if len(val_c) >= 4 and val_c in _universe_compact_set(universe):
        return True
    # Composed-string acceptance. `val` is already norm_string'd (punctuation
    # collapsed to spaces), so split() is the correct tokenizer here.
    toks_all = val.split()
    # Strict pass: for short phrases like "4 miles" or "District 6", the
    # len>=2 filter drops the single-char token below the min-count. Accept
    # when every raw token (single chars included) is in the universe.
    if len(toks_all) >= 2 and all(t in universe for t in toks_all):
        return True
    toks = [t for t in toks_all if len(t) >= 2]
    if len(toks) < 2:
        return False
    compact_universe = _universe_compact_set(universe)
    token_universe = _universe_token_set(universe)
    matched = 0
    for t in toks:
        if t in universe:
            matched += 1
        elif len(t) >= 4 and t in compact_universe:
            matched += 1
        elif len(t) >= 3 and t in token_universe:
            # Token appears as a whitespace-separated piece of some
            # universe entry (e.g. "350" from "f 350"). Required for
            # values like "Ford F-350" where the generator stores make and
            # model as separate fields but the model emits them
            # concatenated.
            matched += 1
    return (matched / len(toks)) >= 0.80


_SMALL_NUMBER_ABS_TOL = 0.5  # counts / single-digit integers → require exact match


def number_in_universe(val: float, numbers: list[float]) -> bool:
    """Allow 1% tolerance for large values; require exact match for small.

    The 1% tolerance was designed for dollar amounts where ±1% is within
    rounding noise. For small integers (counts, IDs, percentages, years)
    the 1% band collapses to effectively zero, but 1% of a large universe
    value can still swallow a small-integer hallucination. Floor it: for
    |val| < 50 (audit 2026-04-17), require an exact numeric match.
    """
    if val == 0:
        return any(n == 0 for n in numbers)
    if abs(val) < 50:
        return any(abs(val - n) < _SMALL_NUMBER_ABS_TOL for n in numbers)
    for n in numbers:
        if n == 0:
            continue
        if abs(val - n) / abs(n) <= 0.01:
            return True
    return False


# ── Extraction walker ─────────────────────────────────────────────────


# Leaves matched by name - these are schema enums or template metadata
# that ground truth doesn't model, so absence from the universe doesn't
# imply hallucination.
SKIP_LEAF_NAMES = {
    # Schema enums
    "coverage", "coverage_type", "coverage_code", "category", "priority",
    "status", "document_type", "statement_type", "period_type",
    "risk_level", "insurability", "entity_type", "construction",
    "occupancy", "mvr_status", "sex", "license_state",
    "risk_grade", "grading_scale", "fiscal_year", "period_end_date",
    "audited", "preparer", "report_type",
    # ACORD form metadata
    "form_edition", "form_title", "form_type", "form_code", "form_number",
    # Producer info - not tracked in GT
    "producer_code", "producer_name", "producer_address", "producer_id",
    # Dates that can be extracted but aren't in GT for many doc types
    "date", "report_date", "issue_date", "inspection_date",
    # Text/summary sections
    "summary", "description", "notes", "remarks",
    "executive_summary", "operations_description", "safety_programs",
    "claims_narrative", "nature_of_business", "loss_history_summary",
    # Enum-like
    "type", "kind", "role", "title",
}

# Path fragments - if the leaf's path contains any of these, skip it
SKIP_PATH_FRAGMENTS = {
    "form_info", "producer",
    "extracted_text_sections", "overall_assessment",
    "recommendations",  # recommendation text is narrative, not structured value
    "risk_highlights",  # free-text positive/concern lists
    # Aggregates - models compute these from line items, GT doesn't
    # necessarily store them. Scored separately by check_arithmetic.
    "grand_totals", "subtotals_by_coverage", "totals",
    "premium_summary", "premium_info",
    "ratios",  # financial ratios are model-computed
}


def _is_aggregate_leaf(leaf: str) -> bool:
    """Paths whose terminal name is an aggregate (model-computed sum)."""
    return leaf.startswith("total_") or leaf.startswith("subtotal_") or leaf in {
        "total", "subtotal", "grand_total", "total_insured_value",
        "total_current_assets", "total_fixed_assets", "total_assets",
        "total_current_liabilities", "total_long_term", "total_liabilities",
        "total_equity", "total_operating_expenses", "total_premium",
        "total_due", "taxes_fees",
    }

# Filler / refusal strings the model uses when it can't extract
FILLER_STRINGS = {
    "multiple", "various", "see below", "see coverages",
    "see above", "see schedule", "see attached", "see page",
    "n a", "na", "tbd", "pending", "unknown", "not applicable",
    "not available", "none", "null",
    # ACORD checkbox schema values: every model emits "checked"/"unchecked"
    # as the literal value for boolean checkbox fields. The rendered page
    # carries a glyph, not the English word, so these were being flagged
    # as fabrications on every ACORD across every model.
    "checked", "unchecked",
}


def walk_extraction(obj: Any, path: tuple = ()):
    if isinstance(obj, dict):
        for k, v in obj.items():
            yield from walk_extraction(v, path + (k,))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            yield from walk_extraction(v, path + (f"[{i}]",))
    else:
        yield path, obj


# ── Per-doc analysis ──────────────────────────────────────────────────


def analyze_doc(extraction: dict, gt_doc: dict, packet_universe: tuple = None) -> dict:
    """Return a hallucination report for one extraction.

    packet_universe: optional (strings_set, numbers_list) pre-computed from
    the entire packet. Shared customer info (insured, producer, preparer,
    carrier) repeats across docs, so matching against the packet-wide
    universe avoids false positives where the model extracts a real
    customer-level value that happens not to be modeled in that specific
    doc's ground truth.
    """
    if packet_universe is not None:
        strings, numbers = packet_universe
    else:
        strings, numbers = collect_universe(gt_doc)

    # Walk extraction, score each leaf
    checked_strings = 0
    hallucinated_strings = 0
    checked_numbers = 0
    hallucinated_numbers = 0
    examples = []  # top offenders

    for path, value in walk_extraction(extraction):
        if value is None or value is True or value is False:
            continue
        leaf = path[-1] if path else ""
        # Normalize the leaf name for matching (drop list indices)
        leaf_bare = leaf if not (isinstance(leaf, str) and leaf.startswith("[")) else (
            path[-2] if len(path) >= 2 else ""
        )
        path_str = ".".join(str(p) for p in path)
        path_parts = {p for p in path if isinstance(p, str) and not p.startswith("[")}

        # Skip template / schema-enum leaves
        if leaf_bare in SKIP_LEAF_NAMES:
            continue
        if path_parts & SKIP_PATH_FRAGMENTS:
            continue
        if _is_aggregate_leaf(leaf_bare):
            continue

        # Numeric-first: numbers coerce cleanly
        num_val = as_float(value)
        is_numeric_field = isinstance(value, (int, float)) and not isinstance(value, bool)
        if num_val is not None and (is_numeric_field or (isinstance(value, str) and _is_mostly_numeric(value))):
            checked_numbers += 1
            if not number_in_universe(num_val, numbers):
                hallucinated_numbers += 1
                examples.append({
                    "kind": "number",
                    "path": path_str,
                    "value": value,
                })
            continue

        # String leaves
        if isinstance(value, str):
            ns = norm_string(value)
            if not looks_meaningful_string(ns):
                continue
            if ns in FILLER_STRINGS:
                continue
            # Skip any "see X" type filler
            if ns.startswith("see "):
                continue
            checked_strings += 1
            if not string_in_universe(ns, strings):
                hallucinated_strings += 1
                # Flag high-value identifiers as high-severity
                severity = "high" if leaf_bare in {
                    "claim_number", "policy_number", "claimant",
                    "license_number", "naics_code", "sic_code",
                } else "normal"
                examples.append({
                    "kind": "string",
                    "path": path_str,
                    "value": value[:120],
                    "severity": severity,
                })

    # Overcount check: extracted list longer than GT list for major lists
    overcount = {}
    for list_key in ["locations", "claims", "drivers", "locations_inspected",
                     "recommendations", "coverages", "forms_attached",
                     "endorsements"]:
        gt_list = gt_doc.get(list_key)
        ext_list = extraction.get(list_key)
        if not isinstance(gt_list, list) or not isinstance(ext_list, list):
            continue
        if len(ext_list) > len(gt_list):
            overcount[list_key] = {
                "gt_count": len(gt_list),
                "ext_count": len(ext_list),
                "excess": len(ext_list) - len(gt_list),
            }

    # Internal arithmetic consistency
    arith_errors = check_arithmetic(extraction)

    return {
        "strings_checked": checked_strings,
        "strings_hallucinated": hallucinated_strings,
        "numbers_checked": checked_numbers,
        "numbers_hallucinated": hallucinated_numbers,
        # Rate is None (not 0.0) when nothing was evaluated, so a doc that
        # the scorer couldn't score on isn't silently reported as "zero
        # hallucinations." Downstream consumers must treat None as "no data"
        # and either filter or surface it explicitly.
        "string_hallucination_rate": (
            hallucinated_strings / checked_strings if checked_strings else None
        ),
        "number_hallucination_rate": (
            hallucinated_numbers / checked_numbers if checked_numbers else None
        ),
        "overcount_lists": overcount,
        "arithmetic_errors": arith_errors,
        "examples": examples,
    }


def _is_mostly_numeric(s: str) -> bool:
    """True when the string is really a formatted number, not prose."""
    s = s.strip()
    return bool(re.fullmatch(r"[\$\s]*-?[\d,]+(\.\d+)?\s*%?", s))


def check_arithmetic(extraction: dict) -> list[dict]:
    """Cheap internal consistency checks."""
    errors = []

    # SOV: compare the extraction's own reported TIV total against the sum
    # of its extracted per-location values. This is an internal-consistency
    # check - both sides come from the extraction, not ground truth.
    locs = extraction.get("locations")
    totals = extraction.get("totals", {}) or {}
    if isinstance(locs, list) and locs:
        reported_tiv = as_float(totals.get("tiv"))
        computed = 0.0
        ok = True
        for loc in locs:
            bv = as_float(loc.get("building_value") if isinstance(loc, dict) else None)
            cv = as_float(loc.get("contents_value") if isinstance(loc, dict) else None)
            bi = as_float(loc.get("bi_value") if isinstance(loc, dict) else None)
            tiv_loc = as_float(loc.get("tiv") if isinstance(loc, dict) else None)
            if tiv_loc is not None:
                computed += tiv_loc
            else:
                parts = [x for x in (bv, cv, bi) if x is not None]
                if parts:
                    computed += sum(parts)
                else:
                    ok = False
                    break
        if ok and reported_tiv is not None and reported_tiv > 0 and computed > 0:
            if abs(computed - reported_tiv) / reported_tiv > 0.02:
                errors.append({
                    "kind": "tiv_sum_mismatch",
                    "reported": reported_tiv,
                    "computed": computed,
                })

    # Loss run: sum of claim incurred vs grand_totals.incurred
    claims = extraction.get("claims")
    gtots = extraction.get("grand_totals", {}) or {}
    if isinstance(claims, list) and claims:
        reported = as_float(gtots.get("incurred"))
        computed = 0.0
        ok = True
        for c in claims:
            if not isinstance(c, dict):
                continue
            inc = as_float(c.get("incurred"))
            if inc is None:
                p = as_float(c.get("paid"))
                r = as_float(c.get("reserved"))
                if p is not None and r is not None:
                    computed += p + r
                else:
                    ok = False
                    break
            else:
                computed += inc
        if ok and reported is not None and reported > 0 and computed > 0:
            if abs(computed - reported) / reported > 0.02:
                errors.append({
                    "kind": "incurred_sum_mismatch",
                    "reported": reported,
                    "computed": computed,
                })

    return errors


# ── Aggregation ───────────────────────────────────────────────────────


def run_model(model_dir: Path, gt_dir: Path) -> dict:
    # Clear the compact-set memoization cache before each model run. The
    # cache keys on id(universe); Python reuses object ids after GC, so a
    # universe set from a prior packet can collide with the id of a fresh
    # universe set and return a stale compact form. We also clear per-packet
    # below - model-level clear catches any leak across model runs.
    _UNIVERSE_COMPACT_CACHE.clear()

    model_report = {
        "model": model_dir.name,
        "docs": {},
        "aggregate": {},
    }

    # category bucketing
    buckets = defaultdict(lambda: {
        "docs": 0,
        "strings_checked": 0, "strings_hallucinated": 0,
        "numbers_checked": 0, "numbers_hallucinated": 0,
        "overcount_docs": 0, "overcount_total_excess": 0,
        "arith_error_docs": 0,
        "high_severity_examples": [],
    })
    difficulty = defaultdict(lambda: {
        "docs": 0,
        "strings_checked": 0, "strings_hallucinated": 0,
        "numbers_checked": 0, "numbers_hallucinated": 0,
    })
    # Track missing/errored extractions at the model level so downstream
    # consumers can see the denominator without guessing. Silently dropping
    # these was a v1 survivor-bias pattern.
    missing_docs = 0
    errored_docs = 0

    for gt_file in sorted(gt_dir.glob("*.json")):
        if gt_file.name.endswith("_summary.json"):
            continue
        gt_data = json.loads(gt_file.read_text())
        packet_id = gt_data.get("packet_id", gt_file.stem)
        diff = gt_data.get("difficulty", "unknown")

        # Build packet-wide universe: packet GT + generator source-of-truth
        # + rendered document content. Shared customer info and fields the
        # packet GT schema doesn't model (agency_customer_id, distances,
        # fire district, etc.) would otherwise register as hallucinations.
        # Clear the compact-form cache before building the new universe so
        # an id() collision with the just-freed universe can't return stale
        # data (see _universe_compact_set).
        _UNIVERSE_COMPACT_CACHE.clear()
        packet_universe = build_packet_universe(gt_data, gt_file)

        for doc_key, doc_gt in gt_data.get("documents", {}).items():
            ext_path = model_dir / f"extraction_{packet_id}_{doc_key}.json"
            if not ext_path.exists():
                missing_docs += 1
                continue
            try:
                ext = json.loads(ext_path.read_text())
            except Exception:
                errored_docs += 1
                continue
            if isinstance(ext, dict) and "error" in ext and set(ext.keys()) <= {
                "error", "packet_id", "doc_type"
            }:
                errored_docs += 1
                continue  # stub

            rep = analyze_doc(ext, doc_gt, packet_universe=packet_universe)
            model_report["docs"][f"{packet_id}/{doc_key}"] = rep

            cat = doc_gt.get("category", "other")
            b = buckets[cat]
            b["docs"] += 1
            b["strings_checked"] += rep["strings_checked"]
            b["strings_hallucinated"] += rep["strings_hallucinated"]
            b["numbers_checked"] += rep["numbers_checked"]
            b["numbers_hallucinated"] += rep["numbers_hallucinated"]
            if rep["overcount_lists"]:
                b["overcount_docs"] += 1
                b["overcount_total_excess"] += sum(
                    v["excess"] for v in rep["overcount_lists"].values()
                )
            if rep["arithmetic_errors"]:
                b["arith_error_docs"] += 1
            for ex in rep["examples"]:
                if ex.get("severity") == "high":
                    b["high_severity_examples"].append({
                        "doc": f"{packet_id}/{doc_key}",
                        **ex,
                    })

            d = difficulty[diff]
            d["docs"] += 1
            d["strings_checked"] += rep["strings_checked"]
            d["strings_hallucinated"] += rep["strings_hallucinated"]
            d["numbers_checked"] += rep["numbers_checked"]
            d["numbers_hallucinated"] += rep["numbers_hallucinated"]

    # Finalize aggregates. Rates are None when nothing was checked, so the
    # output JSON distinguishes "genuine 0% hallucination" from "scorer
    # couldn't evaluate this bucket." Downstream must treat None explicitly.
    agg_cats = {}
    for cat, b in buckets.items():
        agg_cats[cat] = {
            **b,
            "string_hallucination_rate": (
                b["strings_hallucinated"] / b["strings_checked"]
                if b["strings_checked"] else None
            ),
            "number_hallucination_rate": (
                b["numbers_hallucinated"] / b["numbers_checked"]
                if b["numbers_checked"] else None
            ),
            # Keep only 5 examples per category in the summary to avoid giant output
            "high_severity_examples": b["high_severity_examples"][:5],
        }

    agg_diff = {}
    for diff_key, d in difficulty.items():
        agg_diff[diff_key] = {
            **d,
            "string_hallucination_rate": (
                d["strings_hallucinated"] / d["strings_checked"]
                if d["strings_checked"] else None
            ),
            "number_hallucination_rate": (
                d["numbers_hallucinated"] / d["numbers_checked"]
                if d["numbers_checked"] else None
            ),
        }

    all_strings_checked = sum(b["strings_checked"] for b in buckets.values())
    all_strings_hall = sum(b["strings_hallucinated"] for b in buckets.values())
    all_nums_checked = sum(b["numbers_checked"] for b in buckets.values())
    all_nums_hall = sum(b["numbers_hallucinated"] for b in buckets.values())

    model_report["aggregate"] = {
        "overall": {
            "docs": sum(b["docs"] for b in buckets.values()),
            "docs_missing": missing_docs,
            "docs_errored": errored_docs,
            "strings_checked": all_strings_checked,
            "strings_hallucinated": all_strings_hall,
            "string_hallucination_rate": (
                all_strings_hall / all_strings_checked if all_strings_checked else None
            ),
            "numbers_checked": all_nums_checked,
            "numbers_hallucinated": all_nums_hall,
            "number_hallucination_rate": (
                all_nums_hall / all_nums_checked if all_nums_checked else None
            ),
            "overcount_docs": sum(b["overcount_docs"] for b in buckets.values()),
            "arith_error_docs": sum(b["arith_error_docs"] for b in buckets.values()),
        },
        "by_category": agg_cats,
        "by_difficulty": agg_diff,
    }

    return model_report


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ground-truth", type=Path, default=Path("ground_truth"))
    ap.add_argument("--results", type=Path, default=Path("results"))
    ap.add_argument("--output", type=Path, default=Path("results/hallucination_report.json"))
    # Default is the five frontier models being published in the benchmark.
    # A warning is printed at run time listing which dirs will be scored;
    # confirm before shipping any published number.
    ap.add_argument("--models", nargs="*", default=["gpt55", "gpt54", "opus47", "sonnet", "gemini_pro"])
    args = ap.parse_args()

    print("=" * 80)
    print("MODELS BEING SCORED - verify before publishing any number:")
    for m in args.models:
        mdir = args.results / m
        status = "OK" if mdir.is_dir() else "MISSING"
        print(f"  results/{m}  [{status}]")
    print("=" * 80)

    all_reports = {}
    for model in args.models:
        model_dir = args.results / model
        if not model_dir.is_dir():
            print(f"skipping {model}: no dir")
            continue
        print(f"analyzing {model}...")
        all_reports[model] = run_model(model_dir, args.ground_truth)

    args.output.write_text(json.dumps(all_reports, indent=2))
    print(f"\nReport written: {args.output}")

    if _PDF_INGEST_FAILURES:
        print(f"\n⚠  {len(_PDF_INGEST_FAILURES)} PDFs produced no text via pdftotext or OCR.")
        print("   Hallucination rates for these docs depend on a reduced universe.")
        print("   First 5:")
        for p in _PDF_INGEST_FAILURES[:5]:
            print(f"     - {p}")

    # Printable summary
    print("\n" + "=" * 80)
    print("HALLUCINATION SUMMARY")
    print("-" * 80)
    print(f"{'Model':<20} {'Str hall%':>10} {'Num hall%':>10} {'Overcount':>10} {'ArithErr':>10} {'Miss/Err':>10}")

    def _fmt_rate(r: float | None) -> str:
        # Match the float formatter's column width so None rows don't break
        # table alignment.
        return f"{'(no data)':>10}" if r is None else f"{r:>10.1%}"

    for m, rep in all_reports.items():
        o = rep["aggregate"]["overall"]
        print(f"{m:<20} "
              f"{_fmt_rate(o['string_hallucination_rate'])} "
              f"{_fmt_rate(o['number_hallucination_rate'])} "
              f"{o['overcount_docs']:>10} "
              f"{o['arith_error_docs']:>10} "
              f"{o.get('docs_missing', 0) + o.get('docs_errored', 0):>10}")


if __name__ == "__main__":
    main()
