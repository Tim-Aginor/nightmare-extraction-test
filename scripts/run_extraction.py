#!/usr/bin/env python3
"""Multi-model extraction runner for Nightmare Extraction Test.

Supports OpenAI, Anthropic, and Google models.
Tracks token usage and cost per document.

Usage:
    # Run all 4 blog models on all packets
    python scripts/run_extraction.py --ground-truth ground_truth/ --model all

    # Run a specific model
    python scripts/run_extraction.py --ground-truth ground_truth/ --model gpt4o

    # List available models
    python scripts/run_extraction.py --list-models
"""

import argparse
import base64
import json
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass, asdict
from pathlib import Path

import yaml


SCRIPT_DIR = Path(__file__).resolve().parent
# PUBLIC_DIR is the release tree (scripts/, configs/, prompts/, examples/).
# Ground-truth and results paths are resolved from the invocation directory
# via CLI args so the script works whether invoked from inside a workspace
# with all 5 packet GTs or from a user reproducing against
# public/examples/baseline_N1/.
PUBLIC_DIR = SCRIPT_DIR.parent
DEFAULT_CONFIG = PUBLIC_DIR / "configs" / "models.yaml"
SCHEMA_DIR = PUBLIC_DIR / "schemas"


@dataclass
class ExtractionResult:
    """Result of a single document extraction."""
    packet_id: str
    doc_type: str
    model: str
    success: bool
    error: str | None = None
    input_tokens: int = 0
    output_tokens: int = 0
    cost_usd: float = 0.0
    elapsed_s: float = 0.0


# ── Schema loading & per-provider strict-mode adapters ────────────────


def get_schema_for_doc_type(doc_type: str) -> dict | None:
    """Load the JSON Schema for a doc_type, or None if no per-doc schema.

    Mirrors get_prompt_for_doc_type's mapping so every doc that has a prompt
    has a strict schema. ACORD forms map to per-form schemas (§1-B); excel/
    csv variants share the parent schema since the model output shape is the
    same across source formats.
    """
    schema_map = {
        "sov": "sov.schema.json",
        "sov_excel": "sov.schema.json",
        "loss_run": "loss_run.schema.json",
        "loss_run_excel": "loss_run.schema.json",
        "loss_run_csv": "loss_run.schema.json",
        "engineering_report": "engineering_report.schema.json",
        "dec_page": "dec_page.schema.json",
        "driver_schedule": "driver_schedule.schema.json",
        "driver_schedule_excel": "driver_schedule.schema.json",
        "financial_statement": "financial_statement.schema.json",
        "financial_statement_excel": "financial_statement.schema.json",
        "broker_narrative": "narrative.schema.json",
        "supplemental_app": "narrative.schema.json",
        "supplemental_app_trucking": "narrative.schema.json",
        "policy_form": "narrative.schema.json",
        "hybrid_workbook": "narrative.schema.json",
        "supplemental_schedule_excel": "narrative.schema.json",
        "experience_mod": "narrative.schema.json",
        "experience_mod_excel": "narrative.schema.json",
    }
    if doc_type.startswith("acord_"):
        # acord_125 → acord_125.schema.json. Falls back to None if a new
        # ACORD form ships without a per-form schema (would surface as a
        # plain non-strict call, easier to debug than a silent error).
        candidate = SCHEMA_DIR / f"{doc_type}.schema.json"
        if candidate.exists():
            return json.loads(candidate.read_text())
        return None

    fname = schema_map.get(doc_type)
    if not fname:
        return None
    path = SCHEMA_DIR / fname
    if not path.exists():
        return None
    return json.loads(path.read_text())


def _inject_property_ordering(schema: dict) -> dict:
    """Recursively add `propertyOrdering` to every object in the schema.

    Gemini reorders properties alphabetically by default, which would create
    spurious cross-model output diffs. Per the plan §1, propertyOrdering is
    mandatory on Gemini calls.
    """
    import copy
    out = copy.deepcopy(schema)

    def walk(node):
        if isinstance(node, dict):
            if node.get("type") == "object" and isinstance(node.get("properties"), dict):
                node["propertyOrdering"] = list(node["properties"].keys())
            for v in node.values():
                walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    walk(out)
    return out


# ── Provider: OpenAI ───────────────────────────────────────────────────


def extract_openai(doc_path: Path, doc_type: str, model: str, prompt: str,
                   reasoning: dict | None = None,
                   schema: dict | None = None) -> tuple[dict, int, int]:
    """Extract using OpenAI API. Output cap is fixed at 32000 tokens.

    reasoning: optional dict like {"effort": "high"} or {"effort": "xhigh"}.
    When None (default), calls the model in its default mode - same as the
    original v1 benchmark run.
    """
    import openai
    client = openai.OpenAI()

    ext = doc_path.suffix.lower()

    if ext == ".pdf":
        pdf_b64 = base64.standard_b64encode(doc_path.read_bytes()).decode()
        messages = [{
            "role": "user",
            "content": [
                {"type": "file", "file": {
                    "filename": doc_path.name,
                    "file_data": f"data:application/pdf;base64,{pdf_b64}",
                }},
                {"type": "text", "text": prompt},
            ],
        }]
    elif ext == ".xlsx":
        text_repr = excel_to_text(doc_path)
        messages = [{"role": "user", "content": f"Excel contents:\n\n{text_repr}\n\n{prompt}"}]
    elif ext == ".csv":
        csv_text = read_csv_with_fallback(doc_path)
        messages = [{"role": "user", "content": f"CSV contents:\n\n{csv_text}\n\n{prompt}"}]
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    kwargs = {
        "model": model,
        "messages": messages,
        "max_completion_tokens": 32000,
    }
    if reasoning and "effort" in reasoning:
        kwargs["reasoning_effort"] = reasoning["effort"]
        # 128K matches the Anthropic reasoning ceiling. Pre-2026-05-11 we used
        # 64K here while Anthropic ran at 128K — an asymmetric cap. Anthropic's
        # observed max was ~37K, so neither cap was actually hit on the v1.0
        # corpus, but the asymmetry was visible in the audit.
        kwargs["max_completion_tokens"] = 128000

    if schema is not None:
        # Strict structured outputs. The schema must satisfy OpenAI's strict
        # subset: additionalProperties: false everywhere + every property in
        # required (both enforced by _build_schemas.py).
        kwargs["response_format"] = {
            "type": "json_schema",
            "json_schema": {
                "name": schema.get("title", "extraction"),
                "strict": True,
                "schema": {k: v for k, v in schema.items()
                           if k not in ("$schema", "title")},
            },
        }

    # Retry parity with the Anthropic / Gemini paths. Pre-2026-05-11 OpenAI
    # had no retry loop — a transient 429/5xx became a permanent error stub
    # while the same transient on Anthropic was retried 5 times. 5 attempts
    # with exp backoff capped at 60s matches the Anthropic budget.
    last_err = None
    for attempt in range(5):
        try:
            response = client.chat.completions.create(**kwargs)
            text = response.choices[0].message.content
            parsed = parse_json_response(text)
            return parsed, response.usage.prompt_tokens, response.usage.completion_tokens
        except (openai.RateLimitError, openai.APITimeoutError,
                openai.APIConnectionError, openai.InternalServerError) as e:
            last_err = e
            time.sleep(min(2 ** attempt * 3, 60))
    raise last_err


# ── Provider: Anthropic ────────────────────────────────────────────────


def extract_anthropic(doc_path: Path, doc_type: str, model: str, prompt: str,
                      reasoning: dict | None = None,
                      schema: dict | None = None) -> tuple[dict, int, int]:
    """Extract using Anthropic API. Output cap is 32000 by default; raised
    when extended thinking is enabled.

    reasoning: optional dict like {"effort": "high"}. Supported values
    differ by model: Opus 4.7 accepts low/medium/high/xhigh/max; Sonnet 4.6
    only accepts low/medium/high/max (xhigh is rejected with a 400).
    When None, calls the model in default mode - same as v1.

    API shape (verified against Opus 4.7 on 2026-04-21): the new models
    require thinking={"type": "adaptive"} plus output_config={"effort": ...}.
    The old thinking={"type": "enabled", "budget_tokens": N} is rejected
    with "not supported for this model".
    """
    import anthropic
    # 600s wasn't enough for Sonnet 4.6 high-effort on N4/N5 loss_runs - those
    # deterministically time out. 1200s covers the slowest observed cases.
    _override = os.environ.get("ANTHROPIC_API_TIMEOUT")
    _timeout = float(_override) if _override else (1200.0 if reasoning else 300.0)
    _retries = int(os.environ.get("ANTHROPIC_MAX_RETRIES", "2"))
    client = anthropic.Anthropic(timeout=_timeout, max_retries=_retries)

    ext = doc_path.suffix.lower()

    if ext == ".pdf":
        pdf_b64 = base64.standard_b64encode(doc_path.read_bytes()).decode()
        content = [
            {"type": "document", "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": pdf_b64,
            }},
            {"type": "text", "text": prompt},
        ]
    elif ext == ".xlsx":
        text_repr = excel_to_text(doc_path)
        content = [{"type": "text", "text": f"Excel contents:\n\n{text_repr}\n\n{prompt}"}]
    elif ext == ".csv":
        csv_text = read_csv_with_fallback(doc_path)
        content = [{"type": "text", "text": f"CSV contents:\n\n{csv_text}\n\n{prompt}"}]
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    kwargs = {
        "model": model,
        "max_tokens": 32000,
        "messages": [{"role": "user", "content": content}],
        # Anthropic API exposes no `seed` parameter as of 2026-05; extended
        # thinking also requires temperature=1.0. Note: v1 methodology
        # dropped seed pinning across all providers — no `seed=` kwarg is
        # passed on the OpenAI or Gemini paths either, so all three are
        # equally unseeded. Per-doc rate noise falls inside the bootstrap
        # CIs reported in paired_stats.
    }
    if reasoning and "effort" in reasoning:
        # New adaptive mode: Anthropic sizes the thinking budget itself based
        # on the effort level. No budget_tokens parameter needed.
        kwargs["thinking"] = {"type": "adaptive"}
        kwargs["output_config"] = {"effort": reasoning["effort"]}
        # Give the model enough output room for thinking + answer. 64K wasn't
        # enough for Sonnet high on N5 loss_runs - thinking consumed the cap
        # and left an empty or truncated text block. 128K covers the slowest
        # observed cases.
        kwargs["max_tokens"] = 128000

    if schema is not None:
        # Anthropic enforces JSON shape via tool_use. Define a single tool
        # whose input_schema is our JSON schema. With one tool defined, the
        # model reliably calls it under tool_choice="auto".
        # NOTE: Anthropic rejects {thinking + tool_choice:tool|any} with
        # "Thinking may not be enabled when tool_choice forces tool use."
        # so we use "tool" only when thinking is off, "auto" otherwise.
        tool_input_schema = {k: v for k, v in schema.items()
                             if k not in ("$schema", "title")}
        kwargs["tools"] = [{
            "name": "extract",
            "description": "Return the extracted document fields.",
            "input_schema": tool_input_schema,
        }]
        if "thinking" in kwargs:
            kwargs["tool_choice"] = {"type": "auto"}
        else:
            kwargs["tool_choice"] = {"type": "tool", "name": "extract"}

    # Retry on rate limits / overloaded. Parallel runs against one API key
    # hit 429/529 regularly; exponential backoff handles the transient cases.
    last_err = None
    for attempt in range(5):
        try:
            response = client.messages.create(**kwargs)
            # When schema is wired the response carries a tool_use block whose
            # `.input` is already a parsed dict. Otherwise fall back to the
            # text-block path. With extended thinking on, response.content also
            # contains a ThinkingBlock - skip past it via type matching.
            tool_block = next((b for b in response.content
                               if getattr(b, "type", None) == "tool_use"), None)
            if tool_block is not None:
                parsed = tool_block.input
            else:
                text = next((b.text for b in response.content
                             if getattr(b, "type", None) == "text"), "")
                parsed = parse_json_response(text)
            # Defensive unwrap: Anthropic's tool_use input_schema validation
            # is best-effort, not strict. Opus 4.7 wraps the payload in an
            # envelope key on a non-trivial fraction of docs. Observed keys
            # span {data, input, extract, document, extracted_data, ...} —
            # an allowlist kept growing per run, so detect by schema shape:
            # if the only top-level key is NOT a schema property and the
            # inner dict has ≥2 keys that ARE schema properties, unwrap.
            if (schema is not None and isinstance(parsed, dict)
                    and len(parsed) == 1):
                only_key = next(iter(parsed))
                schema_props = set(schema.get("properties", {}).keys())
                inner = parsed[only_key]
                if (only_key not in schema_props
                        and isinstance(inner, dict)
                        and schema_props
                        and len(set(inner.keys()) & schema_props) >= 2):
                    parsed = inner
            return parsed, response.usage.input_tokens, response.usage.output_tokens
        except (anthropic.RateLimitError, anthropic.APIStatusError,
                anthropic.APIConnectionError) as e:
            last_err = e
            status = getattr(e, "status_code", None)
            if status and status not in (429, 500, 502, 503, 529):
                raise
            time.sleep(min(2 ** attempt * 3, 60))
    raise last_err


# ── Provider: Google ───────────────────────────────────────────────────


def extract_google(doc_path: Path, doc_type: str, model: str, prompt: str,
                   reasoning: dict | None = None,
                   schema: dict | None = None) -> tuple[dict, int, int]:
    """Extract using Google Gemini API via the google-genai SDK.

    reasoning: optional dict. Two shapes supported:
        {"thinking_level": "HIGH"}       - named level (MINIMAL/LOW/MEDIUM/HIGH)
        {"thinking_budget": 32000}       - explicit token budget; -1 = dynamic
    When None, calls the model in default mode.

    The old google-generativeai SDK (0.8.x) is deprecated and does not support
    ThinkingConfig. This uses google-genai >= 1.70.
    """
    from google import genai
    from google.genai import types, errors as genai_errors

    client = genai.Client(api_key=os.environ["GOOGLE_API_KEY"])

    ext = doc_path.suffix.lower()

    if ext == ".pdf":
        pdf_bytes = doc_path.read_bytes()
        contents = [
            types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"),
            prompt,
        ]
    elif ext == ".xlsx":
        text_repr = excel_to_text(doc_path)
        contents = f"Excel contents:\n\n{text_repr}\n\n{prompt}"
    elif ext == ".csv":
        csv_text = read_csv_with_fallback(doc_path)
        contents = f"CSV contents:\n\n{csv_text}\n\n{prompt}"
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    gen_config_kwargs: dict = {
        # Match the OpenAI/Anthropic 128K ceiling for symmetry. Gemini API
        # previously ran uncapped; observed max was well under 128K.
        "max_output_tokens": 128000,
    }
    if reasoning:
        thinking_kwargs = {"include_thoughts": False}
        if "thinking_level" in reasoning:
            thinking_kwargs["thinking_level"] = types.ThinkingLevel(reasoning["thinking_level"])
        if "thinking_budget" in reasoning:
            thinking_kwargs["thinking_budget"] = reasoning["thinking_budget"]
        gen_config_kwargs["thinking_config"] = types.ThinkingConfig(**thinking_kwargs)

    if schema is not None:
        # Gemini structured output. response_json_schema accepts the JSON
        # Schema with propertyOrdering injected on every object - default
        # alphabetical reordering would create false cross-model diffs.
        gen_config_kwargs["response_mime_type"] = "application/json"
        ordered = _inject_property_ordering(
            {k: v for k, v in schema.items() if k not in ("$schema", "title")}
        )
        gen_config_kwargs["response_json_schema"] = ordered

    gen_config = types.GenerateContentConfig(**gen_config_kwargs) if gen_config_kwargs else None

    # Retry on 429/5xx. Mirrors the Anthropic path - specific N1/N2 ACORDs
    # hit 503 "high demand" deterministically on single-shot calls, and the
    # initial 5-attempt/~90s budget wasn't enough. 7 attempts w/ 60s cap gives
    # ~250s of backoff, which covers the observed "high demand" windows.
    last_err = None
    for attempt in range(7):
        try:
            response = client.models.generate_content(
                model=model,
                contents=contents,
                config=gen_config,
            )
            break
        except genai_errors.APIError as e:
            last_err = e
            code = getattr(e, "code", None)
            if code not in (429, 499, 500, 502, 503, 504):
                raise
            time.sleep(min(2 ** attempt * 3, 60))
    else:
        raise last_err

    text = response.text
    usage = response.usage_metadata
    in_tok = getattr(usage, "prompt_token_count", 0) or 0
    out_tok = getattr(usage, "candidates_token_count", 0) or 0
    return parse_json_response(text), in_tok, out_tok


# ── Provider: Claude Code CLI ─────────────────────────────────────────


def extract_claude_code(doc_path: Path, doc_type: str, model: str, prompt: str) -> tuple[dict, int, int]:
    """Extract using Claude Code CLI (`claude -p`). Free on Max plan."""
    import subprocess

    abs_path = str(doc_path.resolve())
    ext = doc_path.suffix.lower()

    if ext == ".pdf":
        full_prompt = (
            f"{prompt}\n\n"
            f"Now read the PDF file at {abs_path} and extract the data. "
            f"Return ONLY the JSON object, no other text."
        )
    elif ext == ".xlsx":
        full_prompt = (
            f"{prompt}\n\n"
            f"The document is an Excel file at {abs_path}. "
            f"Read it using the Read tool (it supports Excel), then extract the data. "
            f"Return ONLY the JSON object, no other text."
        )
    elif ext == ".csv":
        full_prompt = (
            f"{prompt}\n\n"
            f"Now read the CSV file at {abs_path} and extract the data. "
            f"Return ONLY the JSON object, no other text."
        )
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    cmd = [
        "claude", "-p", full_prompt,
        "--model", model,
        "--output-format", "json",
        "--max-turns", "5",
        "--allowedTools", "Read,Bash",
    ]

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=1500, stdin=subprocess.DEVNULL)

    if result.returncode != 0:
        raise RuntimeError(f"claude exited {result.returncode}: {result.stderr[:500]}")

    envelope = json.loads(result.stdout)

    # Extract token usage from JSON envelope
    input_tokens = envelope.get("input_tokens", 0)
    output_tokens = envelope.get("output_tokens", 0)
    if "usage" in envelope:
        input_tokens = envelope["usage"].get("input_tokens", input_tokens)
        output_tokens = envelope["usage"].get("output_tokens", output_tokens)

    text = envelope.get("result", "")
    parsed = parse_json_response(text)
    return parsed, input_tokens, output_tokens


# ── Helpers ────────────────────────────────────────────────────────────


def excel_to_text(doc_path: Path) -> str:
    """Convert Excel file to text representation."""
    import openpyxl
    wb = openpyxl.load_workbook(doc_path, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.sheet_state == "hidden":
            continue
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append("\t".join(str(c) if c is not None else "" for c in row))
        sheets.append(f"=== Sheet: {name} ===\n" + "\n".join(rows))
    return "\n\n".join(sheets)


def read_csv_with_fallback(doc_path: Path) -> str:
    """Read CSV with encoding fallback."""
    for enc in ["utf-8", "utf-8-sig", "windows-1252", "latin-1"]:
        try:
            return doc_path.read_text(encoding=enc)
        except (UnicodeDecodeError, ValueError):
            continue
    return doc_path.read_text(encoding="latin-1", errors="replace")


def parse_json_response(text: str) -> dict:
    """Extract JSON from model response.

    Handles: bare JSON, fenced JSON (```json ... ```), JSON with trailing
    prose, and the case where the model emits multiple concatenated JSON
    objects (returns the first complete one).
    """
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("\n", 1)[1].rsplit("```", 1)[0].strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    # raw_decode returns the first complete JSON value + its end offset, so
    # concatenated objects ("{...}\n{...}") and JSON-then-prose both parse.
    idx = cleaned.find("{")
    if idx >= 0:
        try:
            obj, _ = json.JSONDecoder().raw_decode(cleaned[idx:])
            return obj
        except json.JSONDecodeError:
            pass
    raise ValueError(f"No valid JSON found in response ({len(text)} chars)")


def compute_cost(in_tok: int, out_tok: int, pricing: dict) -> float:
    """Compute USD cost from token counts."""
    if not pricing:
        return 0.0
    return round(
        (in_tok / 1_000_000) * pricing.get("input", 0) +
        (out_tok / 1_000_000) * pricing.get("output", 0),
        6
    )


def get_prompt_for_doc_type(doc_type: str, prompts_config: dict) -> str:
    """Load the appropriate prompt for a document type."""
    # Map doc types to prompt files
    prompt_map = {
        "sov": "sov_extraction.md",
        "sov_excel": "sov_extraction.md",
        "loss_run": "loss_run_extraction.md",
        "loss_run_excel": "loss_run_extraction.md",
        "loss_run_csv": "loss_run_extraction.md",
        "engineering_report": "engineering_report_extraction.md",
        "dec_page": "dec_page_extraction.md",
        "driver_schedule": "driver_schedule_extraction.md",
        "driver_schedule_excel": "driver_schedule_extraction.md",
        "financial_statement": "financial_statement_extraction.md",
        "financial_statement_excel": "financial_statement_extraction.md",
        "broker_narrative": "narrative_extraction.md",
        "supplemental_app": "narrative_extraction.md",
        "supplemental_app_trucking": "narrative_extraction.md",
        "policy_form": "narrative_extraction.md",
        "hybrid_workbook": "narrative_extraction.md",
        "supplemental_schedule_excel": "narrative_extraction.md",
        "experience_mod": "narrative_extraction.md",
        "experience_mod_excel": "narrative_extraction.md",
    }

    # ACORD forms all use the same prompt
    if doc_type.startswith("acord_"):
        prompt_file = "acord_form_extraction.md"
    else:
        prompt_file = prompt_map.get(doc_type, "narrative_extraction.md")

    prompt_path = PUBLIC_DIR / "prompts" / prompt_file
    if not prompt_path.exists():
        prompt_path = PUBLIC_DIR / "prompts" / "narrative_extraction.md"

    return prompt_path.read_text()


# ── Main ───────────────────────────────────────────────────────────────


def process_document(
    doc_path: Path,
    doc_type: str,
    packet_id: str,
    model_config: dict,
    pricing_table: dict,
    prompt: str,
    output_dir: Path,
    schema: dict | None = None,
) -> ExtractionResult:
    """Process a single document with a model."""
    model_name = model_config["name"]
    provider = model_config["provider"]
    model_id = model_config["model"]
    reasoning = model_config.get("reasoning")

    out_path = output_dir / f"extraction_{packet_id}_{doc_type}.json"
    if out_path.exists():
        # Skip only if previous run succeeded. Stub failure writes ({"error":...,
        # "packet_id":..., "doc_type":...}) used to be cached as permanent FAILs
        # because skip-if-exists treated them as "done." That made transient
        # timeouts non-reproducible across machines: the rerun script saw the
        # stub and did nothing, so a fresh clone got different n than the
        # writeup. Audit 2026-05-11.
        try:
            cached = json.loads(out_path.read_text())
            if isinstance(cached, dict) and "error" in cached and set(cached.keys()) <= {
                "error", "packet_id", "doc_type"
            }:
                # Previous attempt was a stub failure; retry this run.
                pass
            else:
                return None  # genuine prior success
        except Exception:
            # Unreadable cache → retry.
            pass

    start = time.time()
    try:
        if provider == "openai":
            parsed, in_tok, out_tok = extract_openai(doc_path, doc_type, model_id, prompt, reasoning, schema=schema)
        elif provider == "anthropic":
            parsed, in_tok, out_tok = extract_anthropic(doc_path, doc_type, model_id, prompt, reasoning, schema=schema)
        elif provider == "google":
            parsed, in_tok, out_tok = extract_google(doc_path, doc_type, model_id, prompt, reasoning, schema=schema)
        elif provider == "claude-code":
            # Claude Code CLI is a free-form text path - no provider strict
            # mode hook. Schema is unused here; the prompt's documented JSON
            # shape is the only contract.
            parsed, in_tok, out_tok = extract_claude_code(doc_path, doc_type, model_id, prompt)
        else:
            raise ValueError(f"Unknown provider: {provider}")

        elapsed = time.time() - start
        pricing_key = model_config.get("pricing_key", model_id)
        cost = compute_cost(in_tok, out_tok, pricing_table.get(pricing_key, {}))

        out_path.write_text(json.dumps(parsed, indent=2))

        return ExtractionResult(
            packet_id=packet_id,
            doc_type=doc_type,
            model=model_name,
            success=True,
            input_tokens=in_tok,
            output_tokens=out_tok,
            cost_usd=cost,
            elapsed_s=elapsed,
        )

    except Exception as e:
        elapsed = time.time() - start
        error_data = {"error": str(e), "packet_id": packet_id, "doc_type": doc_type}
        out_path.write_text(json.dumps(error_data, indent=2))
        return ExtractionResult(
            packet_id=packet_id,
            doc_type=doc_type,
            model=model_name,
            success=False,
            error=str(e),
            elapsed_s=elapsed,
        )


def run_model(model_config: dict, gt_dir: Path, output_dir: Path,
              config: dict) -> dict:
    """Run extraction for one model across all packets."""
    model_name = model_config["name"]
    model_output = output_dir / model_name
    model_output.mkdir(parents=True, exist_ok=True)

    pricing_table = config.get("pricing", {})
    prompts_config = config.get("prompts", {})

    results = []

    print(f"\n{'='*60}")
    print(f"MODEL: {model_name} ({model_config['provider']} / {model_config['model']})")
    print(f"  {model_config.get('description', '')}")
    print(f"{'='*60}")

    for gt_file in sorted(gt_dir.glob("*.json")):
        if gt_file.name.endswith("_summary.csv"):
            continue

        gt_data = json.loads(gt_file.read_text())
        packet_id = gt_data.get("packet_id", gt_file.stem)

        for doc_type, doc_gt in gt_data.get("documents", {}).items():
            doc_path = Path(doc_gt.get("document_path", ""))
            # Relative paths resolve against the GT file's real location so
            # public users who symlink a packet-level GT into ground_truth/
            # still find the sibling documents/ dir.
            if not doc_path.is_absolute():
                doc_path = gt_file.resolve().parent / doc_path
            if not doc_path.exists():
                print(f"  SKIP: {packet_id}/{doc_type} - document not found")
                continue

            prompt = get_prompt_for_doc_type(doc_type, prompts_config)
            schema = get_schema_for_doc_type(doc_type)

            result = process_document(
                doc_path, doc_type, packet_id,
                model_config, pricing_table, prompt, model_output,
                schema=schema,
            )

            if result is None:
                print(f"  SKIP: {packet_id}/{doc_type} - already extracted")
                continue

            results.append(result)
            status = "OK  " if result.success else "FAIL"
            print(f"  {status}: {packet_id}/{doc_type} | "
                  f"{result.input_tokens + result.output_tokens} tok | "
                  f"${result.cost_usd:.4f} | {result.elapsed_s:.1f}s")

    # Save summary
    ok = sum(1 for r in results if r.success)
    fail = sum(1 for r in results if not r.success)
    total_cost = sum(r.cost_usd for r in results)
    total_time = sum(r.elapsed_s for r in results)

    summary = {
        "model": model_name,
        "provider": model_config["provider"],
        "total_docs": len(results),
        "ok": ok,
        "failed": fail,
        "total_cost_usd": round(total_cost, 4),
        "total_time_s": round(total_time, 2),
        "results": [asdict(r) for r in results],
    }

    (model_output / "run_summary.json").write_text(json.dumps(summary, indent=2))

    print(f"\n  Summary: {ok}/{len(results)} ok, ${total_cost:.2f} total, {total_time/60:.1f} min")

    return summary


def main():
    parser = argparse.ArgumentParser(description="Run nightmare benchmark extractions")
    parser.add_argument("--ground-truth", type=Path, default=Path("ground_truth"),
                        help="Ground truth directory (resolved from cwd)")
    parser.add_argument("--output", type=Path, default=Path("results"),
                        help="Output directory for extractions (resolved from cwd)")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG,
                        help="Model config YAML")
    parser.add_argument("--model", type=str, default="all",
                        help="Model to run: a name, or one of "
                             "'all' / 'blog' / 'reasoning' / 'reasoning_high' / "
                             "'reasoning_xhigh' / 'gpt54_sweep' / 'gpt55_sweep' / "
                             "'extended'")
    parser.add_argument("--list-models", action="store_true",
                        help="List available models and exit")
    args = parser.parse_args()

    config = yaml.safe_load(args.config.read_text())

    if args.list_models:
        print(f"\n{'Name':<15} {'Provider':<12} {'Model':<30} {'Description'}")
        print("-" * 85)
        for m in config["models"]:
            ext = " (extended)" if m.get("extended") else ""
            print(f"{m['name']:<15} {m['provider']:<12} {m['model']:<30} {m.get('description', '')}{ext}")
        return

    if not args.ground_truth.exists():
        print(f"ERROR: Ground truth directory not found: {args.ground_truth}")
        print("Run: python scripts/generate_ground_truth.py --generator-output <path>")
        sys.exit(1)

    args.output.mkdir(parents=True, exist_ok=True)

    # Determine which models to run
    if args.model == "all":
        models = [m for m in config["models"] if not m.get("extended")]
    elif args.model == "blog":
        models = [m for m in config["models"] if m.get("blog_order")]
        models.sort(key=lambda m: m.get("blog_order", 99))
    elif args.model == "reasoning":
        models = [m for m in config["models"] if m.get("reasoning")]
    elif args.model == "reasoning_high":
        models = [m for m in config["models"] if m.get("reasoning_level") == "HIGH"]
    elif args.model == "reasoning_xhigh":
        models = [m for m in config["models"] if m.get("reasoning_level") == "XHIGH"]
    elif args.model == "gpt54_sweep":
        # Ofir-matched GPT-5.4 thinking-level sweep: low → medium → high → xhigh
        level_order = {"LOW": 0, "MEDIUM": 1, "HIGH": 2, "XHIGH": 3}
        models = [m for m in config["models"] if m.get("gpt_sweep")]
        models.sort(key=lambda m: level_order.get(m.get("reasoning_level"), 99))
    elif args.model == "gpt55_sweep":
        # GPT-5.5 default + high + xhigh. Default has no reasoning_level,
        # so include it first, then sort the rest by effort.
        level_order = {None: 0, "HIGH": 1, "XHIGH": 2}
        models = [m for m in config["models"]
                  if m.get("model") == "gpt-5.5" and
                  (m.get("gpt55_sweep") or not m.get("reasoning"))]
        models.sort(key=lambda m: level_order.get(m.get("reasoning_level"), 99))
    elif args.model == "extended":
        models = config["models"]
    else:
        models = [m for m in config["models"] if m["name"] == args.model]
        if not models:
            print(f"ERROR: Model '{args.model}' not found")
            sys.exit(1)

    print(f"Running {len(models)} model(s)...")

    all_summaries = []
    for model_config in models:
        summary = run_model(model_config, args.ground_truth, args.output, config)
        all_summaries.append(summary)

    # Final summary
    if len(all_summaries) > 1:
        print(f"\n{'='*60}")
        print("ALL MODELS COMPLETE")
        print(f"{'='*60}")
        for s in all_summaries:
            print(f"  {s['model']:<15} {s['ok']}/{s['total_docs']} ok  "
                  f"${s['total_cost_usd']:.2f}  {s['total_time_s']/60:.1f} min")


if __name__ == "__main__":
    main()
